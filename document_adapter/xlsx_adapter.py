"""XLSX 어댑터: openpyxl 기반. 각 워크시트를 하나의 표로 매핑한다.

- table_index = 워크시트 인덱스(0-based), location = 시트 이름
- 좌표 row/col 은 다른 어댑터와 동일하게 0-based 논리 좌표 (openpyxl 은 1-based)
- 병합 셀: top-left 가 anchor, 나머지는 non-anchor. openpyxl 은 병합 non-anchor
  셀이 읽기전용(MergedCell)이라 set_cell 은 anchor 로 redirect 한다.
- fill_form 은 base 구현이 get_tables/get_cell/set_cell 로 자동 동작한다.
"""
from __future__ import annotations

import datetime
import re
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from .base import (
    CellContent,
    CellOutOfBoundsError,
    DocumentAdapter,
    MergeInfo,
    MergedCellWriteError,
    NotImplementedForFormat,
    TableIndexError,
    TableSchema,
    _has_template,
)

# 시트 이름 제약 (Excel): 31자, 아래 문자 금지, 빈 이름 금지.
_INVALID_SHEET_CHARS = set(r"[]:*?/\\")


def _has_border(cell: Any) -> bool:
    """셀에 테두리가 하나라도 있는지 — markdown 역복원의 표 판별 기준."""
    b = getattr(cell, "border", None)
    if b is None:
        return False
    return any(
        getattr(side, "style", None)
        for side in (b.left, b.right, b.top, b.bottom)
        if side is not None
    )


def _is_code_cell(cell: Any) -> bool:
    font = getattr(cell, "font", None)
    return bool(font and font.name == "Consolas")

TAG_PATTERN = re.compile(r"\{\{\s*(\w+)\s*\}\}")

# Excel 열 너비(문자 단위) → cm 근사. Calibri 11 기준 MDW≈7px, padding 5px.
# px → cm: px/96*2.54.
_PX_PER_CM = 96 / 2.54


def _colwidth_to_cm(chars: float | None) -> float | None:
    if chars is None:
        return None
    px = chars * 7 + 5
    return round(px / _PX_PER_CM, 1)


def _rowheight_to_cm(points: float | None) -> float | None:
    if points is None:
        return None
    return round(points / 72 * 2.54, 1)


def _cell_text(v: Any) -> str:
    """셀 값을 사람이 읽는 텍스트로. 날짜는 시간 0 이면 날짜만 표시."""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        if (v.hour, v.minute, v.second, v.microsecond) == (0, 0, 0, 0):
            return v.date().isoformat()
        return v.isoformat(sep=" ")
    if isinstance(v, datetime.date):
        return v.isoformat()
    return str(v)


# 깔끔한 숫자 문자열만 매칭 (정수 / 천단위콤마 / 소수). 전화·ID·날짜(대시) 제외.
_NUM_RE = re.compile(
    r"^-?\d+$|^-?\d{1,3}(?:,\d{3})+$|^-?\d+\.\d+$|^-?\d{1,3}(?:,\d{3})+\.\d+$"
)


def _maybe_number(s: str) -> int | float | None:
    """금액 등 깔끔한 숫자 문자열을 int/float 로. 아니면 None(문자 유지).

    선행 0 정수(우편번호·사번 등)와 대시 포함(전화·날짜)은 문자로 둔다.
    """
    t = s.strip()
    if not _NUM_RE.match(t):
        return None
    plain = t.replace(",", "")
    intpart = plain.lstrip("-").split(".")[0]
    if len(intpart) > 1 and intpart.startswith("0"):
        return None
    try:
        f = float(plain)
        return int(f) if "." not in plain else f
    except ValueError:
        return None


class XlsxAdapter(DocumentAdapter):
    format = "xlsx"

    def _open(self) -> None:
        # rich_text=True: 셀 안의 run 단위 서식(**굵게** 등)을 CellRichText 로
        # 읽어들인다. 기본값(False)이면 로드 시 평문으로 뭉개져 **재저장할 때
        # 인라인 서식이 사라진다** — 보고서 시트(markdown 렌더)를 편집 도구로
        # 한 번만 건드려도 굵게/기울임이 전부 날아가므로 필수.
        # keep_vba=True: .xlsm 의 매크로 보존 (xlsx 는 무해).
        is_macro = self.path.suffix.lower() in (".xlsm", ".xltm")
        self._wb = load_workbook(
            str(self.path), rich_text=True, keep_vba=is_macro
        )
        # 수식의 *캐시된 계산값* 읽기용(data_only). Excel 이 저장한 파일에만 캐시가
        # 있으므로 lazy 로 로드하고, 없거나 실패하면 False 로 표시.
        self._values_wb: Any = None

    def save(self, path: Path | str | None = None) -> Path:
        target = Path(path) if path else self.path
        self._wb.save(str(target))
        self.path = target
        return target

    def _cached_value(self, ws_title: str, row1: int, col1: int) -> Any:
        """수식 셀의 캐시된 계산값(없으면 None). Excel 저장본에서만 존재."""
        if self._values_wb is None:
            try:
                self._values_wb = load_workbook(str(self.path), data_only=True)
            except Exception:
                self._values_wb = False
        if not self._values_wb:
            return None
        try:
            return self._values_wb[ws_title].cell(row=row1, column=col1).value
        except Exception:
            return None

    def _display_text(self, ws, row0: int, col0: int, raw: Any) -> str:
        """표시 텍스트: 수식이고 캐시 계산값이 있으면 그 값을, 없으면 raw."""
        if isinstance(raw, str) and raw.startswith("="):
            cached = self._cached_value(ws.title, row0 + 1, col0 + 1)
            if cached is not None:
                return _cell_text(cached)
        return _cell_text(raw)

    # ---- helpers ----
    def _ws(self, table_index: int):
        sheets = self._wb.worksheets
        if table_index < 0 or table_index >= len(sheets):
            raise TableIndexError(f"XLSX sheet index {table_index} not found")
        return sheets[table_index]

    @staticmethod
    def _merge_map(ws) -> tuple[dict, dict]:
        """(anchor → span) 와 (covered cell → anchor) 매핑을 0-based 로 반환."""
        anchors: dict[tuple[int, int], tuple[int, int]] = {}
        covered: dict[tuple[int, int], tuple[int, int]] = {}
        for rng in ws.merged_cells.ranges:
            r0, c0 = rng.min_row - 1, rng.min_col - 1
            span = (rng.max_row - rng.min_row + 1, rng.max_col - rng.min_col + 1)
            anchors[(r0, c0)] = span
            for r in range(rng.min_row - 1, rng.max_row):
                for c in range(rng.min_col - 1, rng.max_col):
                    if (r, c) != (r0, c0):
                        covered[(r, c)] = (r0, c0)
        return anchors, covered

    @staticmethod
    def _dims(ws) -> tuple[int, int]:
        return ws.max_row or 0, ws.max_column or 0

    # ---- inspection ----
    def get_placeholders(self) -> list[str]:
        keys: set[str] = set()
        for ws in self._wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        keys.update(TAG_PATTERN.findall(cell.value))
        return sorted(keys)

    def get_tables(self, min_rows: int = 1, min_cols: int = 1,
                   preview_rows: int = 4, max_cell_len: int = 40) -> list[TableSchema]:
        schemas: list[TableSchema] = []
        for idx, ws in enumerate(self._wb.worksheets):
            rows, cols = self._dims(ws)
            if rows < min_rows or cols < min_cols:
                continue
            anchors, covered = self._merge_map(ws)
            visible = min(rows, preview_rows)
            preview: list[list[str | None]] = [
                [None] * cols for _ in range(visible)
            ]
            for r in range(visible):
                for c in range(cols):
                    if (r, c) in covered:
                        continue
                    v = ws.cell(row=r + 1, column=c + 1).value
                    preview[r][c] = self._display_text(ws, r, c, v)[:max_cell_len]
            merges = [MergeInfo(anchor=a, span=s) for a, s in anchors.items()]
            col_widths = [
                _colwidth_to_cm(ws.column_dimensions[get_column_letter(c + 1)].width)
                for c in range(cols)
            ]
            row_heights = [
                _rowheight_to_cm(ws.row_dimensions[r + 1].height)
                for r in range(rows)
            ]
            schemas.append(TableSchema(
                index=idx, rows=rows, cols=cols, preview=preview,
                location=ws.title, merges=merges,
                column_widths_cm=col_widths if any(col_widths) else None,
                row_heights_cm=row_heights if any(row_heights) else None,
            ))
        return schemas

    def get_cell(self, table_index: int, row: int, col: int) -> CellContent:
        ws = self._ws(table_index)
        rows, cols = self._dims(ws)
        if row < 0 or col < 0 or row >= rows or col >= cols:
            raise CellOutOfBoundsError(
                f"cell ({row},{col}) out of bounds ({rows}x{cols})")
        anchors, covered = self._merge_map(ws)
        if (row, col) in covered:
            ar, ac = covered[(row, col)]
            is_anchor, anchor, span = False, (ar, ac), anchors[(ar, ac)]
            v = ws.cell(row=ar + 1, column=ac + 1).value
        else:
            is_anchor, anchor = True, (row, col)
            span = anchors.get((row, col), (1, 1))
            v = ws.cell(row=row + 1, column=col + 1).value
        text = self._display_text(ws, anchor[0], anchor[1], v)
        width_cm = _colwidth_to_cm(
            ws.column_dimensions[get_column_letter(anchor[1] + 1)].width)
        height_cm = _rowheight_to_cm(ws.row_dimensions[anchor[0] + 1].height)
        return CellContent(
            row=row, col=col, text=text, paragraphs=[text],
            is_anchor=is_anchor, anchor=anchor, span=span,
            width_cm=width_cm, height_cm=height_cm, char_count=len(text))

    # ---- editing ----
    def render_template(self, context: dict[str, Any], *,
                        on_missing: str = "blank") -> dict[str, list[str]]:
        report = self._render_report(self.get_placeholders(), context, on_missing)
        for ws in self._wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and _has_template(cell.value):
                        cell.value = self._render_text_block(
                            cell.value, context, on_missing)
        return report

    def _resolve_writable(self, ws, row: int, col: int,
                          allow_merge_redirect: bool) -> tuple[int, int]:
        """병합 non-anchor 좌표면 anchor 로 redirect (openpyxl MergedCell 은 읽기전용)."""
        _, covered = self._merge_map(ws)
        if (row, col) in covered:
            if not allow_merge_redirect:
                ar, ac = covered[(row, col)]
                raise MergedCellWriteError(
                    f"cell ({row},{col}) is part of a merge anchored at "
                    f"({ar},{ac}). Write to the anchor, or pass "
                    f"allow_merge_redirect=True.")
            return covered[(row, col)]
        return row, col

    def set_cell(self, table_index: int, row: int, col: int, value: str,
                 *, allow_merge_redirect: bool = False) -> str:
        ws = self._ws(table_index)
        rows, cols = self._dims(ws)
        if row < 0 or col < 0 or row >= rows or col >= cols:
            raise CellOutOfBoundsError(
                f"cell ({row},{col}) out of bounds ({rows}x{cols})")
        wr, wc = self._resolve_writable(ws, row, col, allow_merge_redirect)
        cell = ws.cell(row=wr + 1, column=wc + 1)
        old = _cell_text(cell.value)
        # 깔끔한 숫자(금액 등)는 숫자로 기록해 Excel 수식/합계가 살아있게.
        # 전화·ID·날짜 등은 _maybe_number 가 None → 문자 유지.
        num = _maybe_number(value)
        cell.value = num if num is not None else value
        return old

    def append_to_cell(self, table_index: int, row: int, col: int, value: str,
                       separator: str = "  ", *,
                       allow_merge_redirect: bool = False) -> str:
        ws = self._ws(table_index)
        wr, wc = self._resolve_writable(ws, row, col, allow_merge_redirect)
        cell = ws.cell(row=wr + 1, column=wc + 1)
        old = "" if cell.value is None else str(cell.value)
        cell.value = f"{old}{separator}{value}" if old else value
        return old

    def append_row(self, table_index: int, values: list[str]) -> None:
        ws = self._ws(table_index)
        ws.append(list(values))

    # ---- 행/열 삽입·삭제 ------------------------------------------------
    #
    # openpyxl 의 insert_rows/delete_rows/insert_cols/delete_cols 는 **셀 값만**
    # 옮긴다. 아래는 실측(openpyxl 3.1.5)으로 확인된 미처리 항목이며 전부 여기서
    # 손으로 보정한다:
    #   · 병합 범위(merged_cells)  — 이동 안 됨 → _shift_merges
    #   · 행 높이 / 열 너비        — 이동 안 됨 → _shift_row_heights / _shift_col_widths
    #   · 삽입된 행·열의 서식      — 비어 있음  → 인접 행/열에서 StyleArray 복사
    # 보정 불가(반환 메시지로 고지):
    #   · 셀 수식의 상대참조       — openpyxl 이 재작성하지 않는다
    #   · 차트 series 참조 범위    — 문자열로 굳어 있어 따라오지 않는다

    @staticmethod
    def _merge_boxes(ws) -> list[tuple[int, int, int, int]]:
        """병합 범위를 (min_row, min_col, max_row, max_col) 1-based 로 스냅샷."""
        return [
            (r.min_row, r.min_col, r.max_row, r.max_col)
            for r in list(ws.merged_cells.ranges)
        ]

    @classmethod
    def _shift_merges(cls, ws, at: int, delta: int, *, axis: str) -> None:
        """삽입/삭제 지점 이후의 병합 범위를 delta 만큼 이동한다.

        경계를 가로지르는 병합은 호출 전에 가드로 거른다. 삭제 시 해당
        행/열만으로 이뤄진 병합은 함께 사라진다.
        """
        boxes = cls._merge_boxes(ws)
        if not boxes:
            return
        for r0, c0, r1, c1 in boxes:
            ws.unmerge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)
        for r0, c0, r1, c1 in boxes:
            if axis == "row":
                if delta < 0 and r0 == at and r1 == at:
                    continue                      # 삭제되는 단일행 병합
                if r0 >= at:
                    r0, r1 = r0 + delta, r1 + delta
                elif r1 >= at:
                    r1 += delta                   # 가드 통과분(끝단 접함)만 확장
            else:
                if delta < 0 and c0 == at and c1 == at:
                    continue
                if c0 >= at:
                    c0, c1 = c0 + delta, c1 + delta
                elif c1 >= at:
                    c1 += delta
            if r1 > r0 or c1 > c0:               # 1x1 로 줄면 병합이 아니다
                ws.merge_cells(
                    start_row=r0, start_column=c0, end_row=r1, end_column=c1
                )

    @staticmethod
    def _shift_row_heights(ws, at: int, delta: int) -> None:
        heights = {
            i: d.height for i, d in ws.row_dimensions.items() if d.height is not None
        }
        if not heights:
            return
        moved: dict[int, float] = {}
        for i, h in heights.items():
            if i < at:
                moved[i] = h
            elif delta < 0 and i == at:
                continue
            else:
                moved[i + delta] = h
        for i in list(ws.row_dimensions):
            del ws.row_dimensions[i]
        for i, h in moved.items():
            ws.row_dimensions[i].height = h

    @staticmethod
    def _shift_col_widths(ws, at: int, delta: int) -> None:
        widths = {
            column_index_from_string(k): d.width
            for k, d in ws.column_dimensions.items()
            if d.width is not None
        }
        if not widths:
            return
        moved: dict[int, float] = {}
        for i, w in widths.items():
            if i < at:
                moved[i] = w
            elif delta < 0 and i == at:
                continue
            else:
                moved[i + delta] = w
        for k in list(ws.column_dimensions):
            del ws.column_dimensions[k]
        for i, w in moved.items():
            ws.column_dimensions[get_column_letter(i)].width = w

    @staticmethod
    def _has_formula(ws) -> bool:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
        return False

    def _guard_row_boundary(self, ws, at: int, *, deleting: bool) -> None:
        """삽입/삭제 지점을 세로 병합이 가로지르면 거절."""
        for r0, c0, r1, c1 in self._merge_boxes(ws):
            crosses = (r0 < at <= r1) if not deleting else (r0 <= at <= r1 and r0 != r1)
            if crosses:
                raise NotImplementedForFormat(
                    f"row {at} crosses a vertical merge {get_column_letter(c0)}{r0}:"
                    f"{get_column_letter(c1)}{r1}; unmerge it first."
                )

    def _guard_col_boundary(self, ws, at: int, *, deleting: bool) -> None:
        for r0, c0, r1, c1 in self._merge_boxes(ws):
            crosses = (c0 < at <= c1) if not deleting else (c0 <= at <= c1 and c0 != c1)
            if crosses:
                raise NotImplementedForFormat(
                    f"column {get_column_letter(at)} crosses a horizontal merge "
                    f"{get_column_letter(c0)}{r0}:{get_column_letter(c1)}{r1}; "
                    f"unmerge it first."
                )

    def insert_row(
        self,
        table_index: int,
        values: list[str],
        at_row: int | None = None,
    ) -> int:
        """시트에 행을 삽입한다 (기존 at_row 행은 아래로 밀림).

        서식은 인접 행(삽입 위치의 기존 행, 맨 끝이면 마지막 행)에서 상속한다.
        반환값은 실제 삽입된 0-based 행 인덱스.
        """
        ws = self._ws(table_index)
        n_rows, n_cols = self._dims(ws)
        if at_row is None:
            at_row = n_rows
        if at_row < 0 or at_row > n_rows:
            raise CellOutOfBoundsError(
                f"at_row {at_row} out of range 0..{n_rows}"
            )
        at1 = at_row + 1
        self._guard_row_boundary(ws, at1, deleting=False)

        # 서식 템플릿: 삽입 위치의 기존 행(맨 끝이면 마지막 행)을 미리 캡처.
        n_cols = max(n_cols, len(values))
        tmpl_row = at1 if at_row < n_rows else max(1, n_rows)
        styles = (
            [copy(ws.cell(row=tmpl_row, column=c + 1)._style) for c in range(n_cols)]
            if n_rows
            else []
        )
        tmpl_height = ws.row_dimensions[tmpl_row].height if n_rows else None

        self._shift_merges(ws, at1, +1, axis="row")
        self._shift_row_heights(ws, at1, +1)
        ws.insert_rows(at1)

        for c in range(n_cols):
            cell = ws.cell(row=at1, column=c + 1)
            if c < len(styles):
                cell._style = copy(styles[c])
            if c < len(values) and values[c] != "":
                num = _maybe_number(str(values[c]))
                cell.value = num if num is not None else values[c]
        if tmpl_height is not None:
            ws.row_dimensions[at1].height = tmpl_height
        return at_row

    def insert_column(
        self,
        table_index: int,
        values: list[str],
        at_col: int | None = None,
    ) -> int:
        """시트에 열을 삽입한다 (기존 at_col 열은 오른쪽으로 밀림)."""
        ws = self._ws(table_index)
        n_rows, n_cols = self._dims(ws)
        if at_col is None:
            at_col = n_cols
        if at_col < 0 or at_col > n_cols:
            raise CellOutOfBoundsError(
                f"at_col {at_col} out of range 0..{n_cols}"
            )
        at1 = at_col + 1
        self._guard_col_boundary(ws, at1, deleting=False)

        n_rows = max(n_rows, len(values))
        tmpl_col = at1 if at_col < n_cols else max(1, n_cols)
        styles = (
            [copy(ws.cell(row=r + 1, column=tmpl_col)._style) for r in range(n_rows)]
            if n_cols
            else []
        )
        tmpl_width = (
            ws.column_dimensions[get_column_letter(tmpl_col)].width if n_cols else None
        )

        self._shift_merges(ws, at1, +1, axis="col")
        self._shift_col_widths(ws, at1, +1)
        ws.insert_cols(at1)

        for r in range(n_rows):
            cell = ws.cell(row=r + 1, column=at1)
            if r < len(styles):
                cell._style = copy(styles[r])
            if r < len(values) and values[r] != "":
                num = _maybe_number(str(values[r]))
                cell.value = num if num is not None else values[r]
        if tmpl_width is not None:
            ws.column_dimensions[get_column_letter(at1)].width = tmpl_width
        return at_col

    def delete_row(self, table_index: int, at_row: int) -> list[str]:
        """행을 삭제하고 삭제된 행의 값을 반환한다 (되돌리기용 스냅샷)."""
        ws = self._ws(table_index)
        n_rows, n_cols = self._dims(ws)
        if at_row < 0 or at_row >= n_rows:
            raise CellOutOfBoundsError(
                f"at_row {at_row} out of range 0..{n_rows - 1}"
            )
        at1 = at_row + 1
        self._guard_row_boundary(ws, at1, deleting=True)
        removed = [
            _cell_text(ws.cell(row=at1, column=c + 1).value) for c in range(n_cols)
        ]
        self._shift_merges(ws, at1, -1, axis="row")
        self._shift_row_heights(ws, at1, -1)
        ws.delete_rows(at1)
        return removed

    def delete_column(self, table_index: int, at_col: int) -> list[str]:
        """열을 삭제하고 삭제된 열의 값을 반환한다."""
        ws = self._ws(table_index)
        n_rows, n_cols = self._dims(ws)
        if at_col < 0 or at_col >= n_cols:
            raise CellOutOfBoundsError(
                f"at_col {at_col} out of range 0..{n_cols - 1}"
            )
        at1 = at_col + 1
        self._guard_col_boundary(ws, at1, deleting=True)
        removed = [
            _cell_text(ws.cell(row=r + 1, column=at1).value) for r in range(n_rows)
        ]
        self._shift_merges(ws, at1, -1, axis="col")
        self._shift_col_widths(ws, at1, -1)
        ws.delete_cols(at1)
        return removed

    def has_formulas(self, table_index: int) -> bool:
        """해당 시트에 수식이 있는지 — 삽입/삭제 후 참조 보정 경고 판단용."""
        return self._has_formula(self._ws(table_index))

    # ---- markdown 시트 (서술형 보고서) ----------------------------------
    #
    # markdown 으로 렌더한 보고서 시트는 헤딩/문단은 A열, 불릿은 B열처럼
    # **비균질 레이아웃**이라 셀 좌표 편집이 위험하다 (LLM 이 빈 A열에 써서
    # 옆 불릿과 겹치는 사고가 실제로 발생). 시트 단위로 통째 재렌더하는 경로를
    # 두어 좌표 추측 자체를 없앤다.

    _MD_BULLET = "• "
    _MD_NUMBER = "- "

    def set_sheet_markdown(self, table_index: int, markdown: str) -> int:
        """시트 내용을 전부 비우고 markdown 을 다시 렌더한다.

        생성 계층(`create_document` 의 sheets[].markdown)과 **동일한 렌더러**를
        쓰므로 서식/레이아웃이 항상 일관된다.

        Returns:
            렌더에 사용된 마지막 행 번호(1-based).
        """
        from .generate.xlsx_writer import render_markdown_sheet

        ws = self._ws(table_index)
        for rng in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(rng))
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        for key in list(ws.column_dimensions):
            del ws.column_dimensions[key]
        for key in list(ws.row_dimensions):
            del ws.row_dimensions[key]
        return render_markdown_sheet(ws, markdown or "")

    def get_sheet_markdown(self, table_index: int) -> str:
        """시트를 markdown 으로 되돌린다 (근사 복원).

        ``set_sheet_markdown`` / 생성 계층이 만든 시트는 렌더 규칙이 정해져
        있어 대부분 원문에 가깝게 복원된다. 그 밖의 일반 표 시트는 각 행이
        파이프 표 한 줄로 나온다.
        """
        ws = self._ws(table_index)
        _, covered = self._merge_map(ws)
        lines: list[str] = []
        code_buf: list[str] = []

        def flush_code() -> None:
            if code_buf:
                lines.extend(["```", *code_buf, "```", ""])
                code_buf.clear()

        for r in range(1, (ws.max_row or 0) + 1):
            cells = [
                ws.cell(row=r, column=c)
                for c in range(1, (ws.max_column or 0) + 1)
                if (r - 1, c - 1) not in covered
            ]
            filled = [c for c in cells if c.value not in (None, "")]

            if not filled:
                first = ws.cell(row=r, column=1)
                # 값 없이 아래 테두리만 있는 행 = 수평선
                if first.border and first.border.bottom and first.border.bottom.style:
                    flush_code()
                    lines.extend(["---", ""])
                continue

            if all(_is_code_cell(c) for c in filled):
                code_buf.append(_cell_text(filled[0].value))
                continue
            flush_code()

            # 파이프 표: A열부터 연속으로 채워지고 **모든 셀에 테두리**가 있어야
            # 한다. 테두리를 조건에 넣지 않으면 "A열 문단 + B열 불릿"처럼 우연히
            # 두 칸이 찬 행을 표로 오인해 원문을 파괴한다.
            if (
                len(filled) >= 2
                and filled[0].column == 1
                and [c.column for c in filled] == list(range(1, len(filled) + 1))
                and all(_has_border(c) for c in filled)
            ):
                lines.append(
                    "| " + " | ".join(_cell_text(c.value) for c in filled) + " |"
                )
                continue

            # 그 외에는 채워진 셀을 **각각** 블록으로 변환한다 (한 행에 서술과
            # 불릿이 섞여 있어도 어느 쪽도 잃지 않는다).
            for cell in filled:
                text = _cell_text(cell.value)
                font = cell.font
                if text.startswith(self._MD_BULLET):
                    lines.append(f"- {text[len(self._MD_BULLET):]}")
                elif text.startswith(self._MD_NUMBER):
                    lines.append(f"1. {text[len(self._MD_NUMBER):]}")
                elif font and font.bold and (font.size or 11) >= 16:
                    lines.extend([f"# {text}", ""])
                elif font and font.bold and (font.size or 11) >= 13:
                    lines.extend([f"## {text}", ""])
                elif font and font.bold:
                    lines.extend([f"### {text}", ""])
                elif font and font.italic:
                    lines.extend([f"> {text}", ""])
                else:
                    lines.extend([text, ""])

        flush_code()
        return "\n".join(lines).strip() + "\n"

    # ---- 시트 관리 ------------------------------------------------------
    def _validate_sheet_name(self, name: str, *, exclude: str | None = None) -> str:
        title = (name or "").strip()
        if not title:
            raise ValueError("sheet name must not be empty. 시트 이름은 비울 수 없습니다.")
        bad = sorted(set(title) & _INVALID_SHEET_CHARS)
        if bad:
            raise ValueError(
                f"sheet name contains forbidden characters {bad}. "
                f"시트 이름에 사용할 수 없는 문자가 있습니다: {bad}"
            )
        title = title[:31]
        existing = [t for t in self._wb.sheetnames if t != exclude]
        if title in existing:
            raise ValueError(
                f"sheet {title!r} already exists. 시트 {title!r} 가 이미 존재합니다."
            )
        return title

    def add_sheet(self, name: str, at_index: int | None = None) -> int:
        """새 시트를 만들고 그 table_index(0-based)를 반환한다."""
        title = self._validate_sheet_name(name)
        ws = self._wb.create_sheet(title=title, index=at_index)
        return self._wb.worksheets.index(ws)

    def rename_sheet(self, table_index: int, new_name: str) -> str:
        """시트 이름을 바꾸고 이전 이름을 반환한다."""
        ws = self._ws(table_index)
        old = ws.title
        ws.title = self._validate_sheet_name(new_name, exclude=old)
        return old

    def delete_sheet(self, table_index: int) -> str:
        """시트를 삭제하고 삭제된 시트명을 반환한다. 마지막 1개는 거절."""
        ws = self._ws(table_index)
        if len(self._wb.worksheets) <= 1:
            raise ValueError(
                "cannot delete the last remaining sheet. "
                "마지막 남은 시트는 삭제할 수 없습니다."
            )
        title = ws.title
        self._wb.remove(ws)
        return title
