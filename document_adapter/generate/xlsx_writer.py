"""sheet spec(dict) → XLSX 결정적 렌더러 (LLM 없음).

edit2docs v0.9 `documents/xlsx_engine.xlsx_from_spec`(Apache-2.0)의 렌더
전략을 dict 입력으로 이식 — NOTICE 참조. YAML 의존을 추가하지 않기 위해
스펙은 파싱 완료된 dict 를 받는다 (JSON 파싱은 도구 레이어 몫).

sheet spec 스키마 (LLM 도구 설명과 1:1):
    sheets: [
      # (1) 데이터 시트 — 표
      { "name": "매출 요약",              # 필수, 시트 탭 이름 (31자 절단)
        "headers": ["분기", "매출(억원)"],
        "rows": [["1분기", 120], ...],     # 숫자는 숫자 타입으로
        "widths": [10, 14],                # 선택, 문자 단위 열 폭
        "number_formats": {"B": "#,##0"},  # 선택, 열문자 → Excel 서식
        "freeze": "A2" },                  # 선택, 틀고정 (기본 "A2")

      # (2) 차트 시트 — 다른 시트의 셀 범위를 참조하는 살아있는 차트
      { "name": "차트",
        "charts": [
          { "type": "column",              # 아래 _CHART_TYPES 참조
            "anchor": "B2",                # 차트 좌상단 셀
            "title": "월별 매출",
            "source_sheet": 0,             # 데이터가 있는 시트 (index 또는 name)
            "data":       {"min_col": 2, "max_col": 3, "min_row": 1, "max_row": 13},
            "categories": {"min_col": 1, "max_col": 1, "min_row": 2, "max_row": 13},
            "titles_from_data": true,      # data 첫 행/열을 계열명으로
            "width_cm": 18, "height_cm": 9,
            "x_axis_title": "월", "y_axis_title": "금액" } ] },

      # (3) 보고서 시트 — markdown 을 셀 + 서식으로 렌더
      { "name": "보고서",
        "markdown": "# 분석 결과\\n\\n- 2분기 매출 증가\\n" }
    ]

한 시트는 `headers`+`rows` / `charts` / `markdown` 중 **최소 하나**를 가져야
한다. 셋을 함께 쓸 수도 있다 (표 위에 차트를 얹는 등).

렌더 규칙: 헤더행 굵게+회색 채움+테두리, 데이터행 테두리, 헤더 틀고정,
`=` 로 시작하는 문자열은 수식으로 통과.

차트는 **시트를 모두 만든 뒤 2-pass 로 추가**한다 — `source_sheet` 가 뒤에
정의된 시트를 가리킬 수 있기 때문이다.

모든 검증 실패는 이중어 ValueError — 호출 레이어의 1회 재시도 계약이
이 메시지를 LLM 리마인더로 사용한다.
"""
from __future__ import annotations

import io
import math
from typing import Any

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import (
    AreaChart,
    BarChart,
    DoughnutChart,
    LineChart,
    PieChart,
    RadarChart,
    Reference,
    ScatterChart,
    Series,
)
from openpyxl.chart.marker import Marker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .markdown_parser import Block, Span, parse_markdown

__all__ = ["xlsx_from_sheets", "render_markdown_sheet", "REPORT_COLS"]

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_MIN_WIDTH = 8
_MAX_WIDTH = 60

# ── 보고서(markdown) 시트 렌더 상수 ──────────────────────────────────
_REPORT_COLS = 6                     # A..F 를 본문 폭으로 사용
_REPORT_GUTTER_WIDTH = 3             # A 열 = 들여쓰기 여백
_REPORT_BODY_WIDTH = 20              # B..F 열 폭
_REPORT_TABLE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
_REPORT_CODE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_MEDIUM_BOTTOM = Border(bottom=Side(style="medium", color="808080"))
_THIN_BOTTOM = Border(bottom=Side(style="thin", color="BFBFBF"))
_QUOTE_BORDER = Border(left=Side(style="thick", color="BFBFBF"))
_QUOTE_FONT_COLOR = "595959"
_HEADING_SIZES = {1: 16, 2: 13, 3: 11}
_LINE_HEIGHT = 15.0                  # 1행 기본 높이(pt)
_MAX_ROW_HEIGHT = 409.0              # Excel 상한

# 차트 종류 → (팩토리, BarChart type, grouping, 마커 사용)
# 이름은 pptx `add_chart` 와 동일하게 유지한다 — LLM 이 포맷별로 헷갈리지 않도록.
_CHART_TYPES: dict[str, dict[str, Any]] = {
    "column":         {"cls": BarChart, "type": "col"},
    "column_stacked": {"cls": BarChart, "type": "col", "grouping": "stacked", "overlap": 100},
    "bar":            {"cls": BarChart, "type": "bar"},
    "bar_stacked":    {"cls": BarChart, "type": "bar", "grouping": "stacked", "overlap": 100},
    "line":           {"cls": LineChart},
    "line_markers":   {"cls": LineChart, "markers": True},
    "pie":            {"cls": PieChart},
    "doughnut":       {"cls": DoughnutChart},
    "area":           {"cls": AreaChart},
    "area_stacked":   {"cls": AreaChart, "grouping": "stacked"},
    "radar":          {"cls": RadarChart},
    "scatter":        {"cls": ScatterChart},
}


def _err(en: str, ko: str) -> ValueError:
    return ValueError(f"{en} {ko}")


def _display_len(text: str) -> int:
    """CJK 를 2칸으로 세는 표시 폭 근사 (열 폭·행 높이 추정용)."""
    return sum(2 if ord(ch) > 0x2E7F else 1 for ch in text)


# ─────────────────────────────────────────────────────────────────────
# 인라인 서식 (Span → 셀 값)
# ─────────────────────────────────────────────────────────────────────
def _spans_text(spans: tuple[Span, ...]) -> str:
    return "".join(s.text for s in spans)


def _spans_to_value(spans: tuple[Span, ...]) -> Any:
    """Span 튜플 → 셀 값.

    인라인 마크업이 하나도 없으면 평문 str 을 그대로 돌려준다 (RichText 는
    파일 크기·호환성 비용이 있으므로 필요할 때만 쓴다). 하나라도 있으면
    openpyxl 3.1+ 의 ``CellRichText`` 로 run 단위 서식을 보존한다.
    """
    text = _spans_text(spans)
    if not any(s.bold or s.italic or s.code for s in spans):
        return text
    parts: list[Any] = []
    for s in spans:
        if not s.text:
            continue
        if s.code:
            parts.append(TextBlock(InlineFont(rFont="Consolas"), s.text))
        elif s.bold or s.italic:
            parts.append(
                TextBlock(InlineFont(b=s.bold or None, i=s.italic or None), s.text)
            )
        else:
            parts.append(s.text)
    return CellRichText(parts) if parts else text


# ─────────────────────────────────────────────────────────────────────
# 보고서(markdown) 시트 렌더러
# ─────────────────────────────────────────────────────────────────────
def _merge_span(ws, row: int, first_col: int, last_col: int) -> None:
    if last_col > first_col:
        ws.merge_cells(
            start_row=row, start_column=first_col, end_row=row, end_column=last_col
        )


def _set_row_height(ws, row: int, text: str, avail_width: int) -> None:
    """줄바꿈된 텍스트의 행 높이를 근사 계산해 지정한다.

    openpyxl 은 자동 맞춤(auto-fit)을 지원하지 않는다 — Excel 이 열 때
    계산해주지도 않으므로(수동 조정 전까지) 여기서 추정해 둔다.
    """
    width = max(1, avail_width)
    lines = max(1, math.ceil(_display_len(text) / width))
    ws.row_dimensions[row].height = min(_MAX_ROW_HEIGHT, lines * _LINE_HEIGHT)


def _render_markdown_sheet(ws, md: str) -> int:
    """markdown → 셀 + 서식. 마지막으로 사용한 행 번호를 반환한다."""
    blocks = parse_markdown(md or "")

    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = _REPORT_GUTTER_WIDTH
    for c in range(2, _REPORT_COLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = _REPORT_BODY_WIDTH

    # 병합 폭(문자) 근사 — 행 높이 추정에 쓴다.
    full_width = _REPORT_GUTTER_WIDTH + _REPORT_BODY_WIDTH * (_REPORT_COLS - 1)
    indent_width = _REPORT_BODY_WIDTH * (_REPORT_COLS - 1)

    r = 1
    for blk in blocks:
        if blk.kind == "heading":
            level = max(1, min(6, blk.level))
            size = _HEADING_SIZES.get(level, 11)
            if level == 1 and r > 1:
                r += 1  # 최상위 제목 앞 여백
            cell = ws.cell(row=r, column=1, value=_spans_to_value(blk.spans))
            cell.font = Font(bold=True, size=size)
            cell.alignment = Alignment(vertical="center")
            if level <= 2:
                cell.border = _MEDIUM_BOTTOM if level == 1 else _THIN_BOTTOM
            _merge_span(ws, r, 1, _REPORT_COLS)
            ws.row_dimensions[r].height = _LINE_HEIGHT + size
            r += 2 if level == 1 else 1

        elif blk.kind == "paragraph":
            text = _spans_text(blk.spans)
            cell = ws.cell(row=r, column=1, value=_spans_to_value(blk.spans))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            _merge_span(ws, r, 1, _REPORT_COLS)
            _set_row_height(ws, r, text, full_width)
            r += 1

        elif blk.kind in ("bullet", "numbered"):
            marker = "• " if blk.kind == "bullet" else "- "
            value = _spans_to_value(blk.spans)
            if isinstance(value, CellRichText):
                value = CellRichText([marker, *list(value)])
            else:
                value = marker + str(value)
            cell = ws.cell(row=r, column=2, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            _merge_span(ws, r, 2, _REPORT_COLS)
            _set_row_height(ws, r, marker + _spans_text(blk.spans), indent_width)
            r += 1

        elif blk.kind == "quote":
            text = _spans_text(blk.spans)
            cell = ws.cell(row=r, column=1, value=text)
            cell.font = Font(italic=True, color=_QUOTE_FONT_COLOR)
            cell.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
            cell.border = _QUOTE_BORDER
            _merge_span(ws, r, 1, _REPORT_COLS)
            _set_row_height(ws, r, text, full_width)
            r += 1

        elif blk.kind == "hr":
            cell = ws.cell(row=r, column=1, value=None)
            cell.border = _THIN_BOTTOM
            _merge_span(ws, r, 1, _REPORT_COLS)
            ws.row_dimensions[r].height = _LINE_HEIGHT / 2
            r += 1

        elif blk.kind == "code":
            for line in blk.lines:
                cell = ws.cell(row=r, column=1, value=line)
                cell.font = Font(name="Consolas", size=10)
                cell.fill = _REPORT_CODE_FILL
                cell.alignment = Alignment(vertical="center")
                _merge_span(ws, r, 1, _REPORT_COLS)
                r += 1

        elif blk.kind == "table":
            for t_idx, trow in enumerate(blk.rows):
                for c_idx, cell_spans in enumerate(trow):
                    cell = ws.cell(
                        row=r, column=c_idx + 1, value=_spans_to_value(cell_spans)
                    )
                    cell.border = _BORDER
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if t_idx == 0:
                        cell.font = Font(bold=True)
                        cell.fill = _REPORT_TABLE_FILL
                r += 1
            r += 1  # 표 뒤 여백

    return max(1, r - 1)


# 편집 계층(XlsxAdapter.set_sheet_markdown)이 같은 렌더 규칙을 재사용하도록
# 공개한다 — 생성과 편집이 다른 레이아웃을 만들면 안 되기 때문.
render_markdown_sheet = _render_markdown_sheet
REPORT_COLS = _REPORT_COLS


# ─────────────────────────────────────────────────────────────────────
# 차트
# ─────────────────────────────────────────────────────────────────────
def _ref_from_dict(
    ws, spec: Any, sheet_label: str, field: str
) -> Reference:
    """{"min_col":.., "min_row":.., "max_col":.., "max_row":..} → Reference."""
    if not isinstance(spec, dict):
        raise _err(
            f"sheet {sheet_label!r}: chart `{field}` must be a mapping "
            f'like {{"min_col": 2, "min_row": 1, "max_row": 10}}.',
            f"시트 {sheet_label!r}: 차트 `{field}` 는 "
            f'{{"min_col": 2, "min_row": 1, "max_row": 10}} 형태의 객체여야 합니다.',
        )
    try:
        min_col = int(spec["min_col"])
        min_row = int(spec["min_row"])
    except (KeyError, TypeError, ValueError) as exc:
        raise _err(
            f"sheet {sheet_label!r}: chart `{field}` needs integer "
            f"`min_col` and `min_row` ({exc}).",
            f"시트 {sheet_label!r}: 차트 `{field}` 에는 정수 "
            f"`min_col`, `min_row` 가 필요합니다 ({exc}).",
        ) from exc
    try:
        max_col = int(spec.get("max_col", min_col))
        max_row = int(spec.get("max_row", min_row))
    except (TypeError, ValueError) as exc:
        raise _err(
            f"sheet {sheet_label!r}: chart `{field}` max_col/max_row must be integers.",
            f"시트 {sheet_label!r}: 차트 `{field}` 의 max_col/max_row 는 정수여야 합니다.",
        ) from exc
    if min_col < 1 or min_row < 1 or max_col < min_col or max_row < min_row:
        raise _err(
            f"sheet {sheet_label!r}: chart `{field}` range is invalid "
            f"(cols {min_col}..{max_col}, rows {min_row}..{max_row}); "
            f"indexes are 1-based and max must be >= min.",
            f"시트 {sheet_label!r}: 차트 `{field}` 범위가 올바르지 않습니다 "
            f"(열 {min_col}..{max_col}, 행 {min_row}..{max_row}); "
            f"인덱스는 1부터 시작하며 max 는 min 이상이어야 합니다.",
        )
    return Reference(
        ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
    )


def _make_chart(kind: str, sheet_label: str):
    conf = _CHART_TYPES.get(kind)
    if conf is None:
        raise _err(
            f"sheet {sheet_label!r}: unknown chart `type` {kind!r} "
            f"(supported: {', '.join(sorted(_CHART_TYPES))}).",
            f"시트 {sheet_label!r}: 지원하지 않는 차트 `type` {kind!r} 입니다 "
            f"(지원: {', '.join(sorted(_CHART_TYPES))}).",
        )
    chart = conf["cls"]()
    if "type" in conf:
        chart.type = conf["type"]
    if "grouping" in conf:
        chart.grouping = conf["grouping"]
    if "overlap" in conf:
        chart.overlap = conf["overlap"]
    return chart, bool(conf.get("markers"))


def _resolve_source_ws(
    wb: Workbook, ws_by_index: list[Any], source: Any, sheet_label: str, default_ws
):
    """`source_sheet` (index | name | 생략) → worksheet."""
    if source is None:
        return default_ws
    if isinstance(source, bool):  # bool 은 int 의 서브클래스라 먼저 거른다
        raise _err(
            f"sheet {sheet_label!r}: chart `source_sheet` must be an index or name.",
            f"시트 {sheet_label!r}: 차트 `source_sheet` 는 인덱스 또는 시트명이어야 합니다.",
        )
    if isinstance(source, int):
        if source < 0 or source >= len(ws_by_index):
            raise _err(
                f"sheet {sheet_label!r}: chart `source_sheet` index {source} "
                f"is out of range (0..{len(ws_by_index) - 1}).",
                f"시트 {sheet_label!r}: 차트 `source_sheet` 인덱스 {source} 가 "
                f"범위를 벗어났습니다 (0..{len(ws_by_index) - 1}).",
            )
        return ws_by_index[source]
    name = str(source)[:31]
    if name not in wb.sheetnames:
        raise _err(
            f"sheet {sheet_label!r}: chart `source_sheet` {name!r} not found "
            f"(available: {', '.join(wb.sheetnames)}).",
            f"시트 {sheet_label!r}: 차트 `source_sheet` {name!r} 를 찾을 수 없습니다 "
            f"(존재하는 시트: {', '.join(wb.sheetnames)}).",
        )
    return wb[name]


def _add_charts(
    wb: Workbook, ws, ws_by_index: list[Any], charts: Any, sheet_label: str
) -> None:
    if not isinstance(charts, list):
        raise _err(
            f"sheet {sheet_label!r}: `charts` must be a list.",
            f"시트 {sheet_label!r}: `charts` 는 리스트여야 합니다.",
        )
    for c_idx, spec in enumerate(charts):
        if not isinstance(spec, dict):
            raise _err(
                f"sheet {sheet_label!r}: charts[{c_idx}] must be a mapping.",
                f"시트 {sheet_label!r}: charts[{c_idx}] 항목은 객체(dict)여야 합니다.",
            )
        kind = str(spec.get("type") or "column").strip().lower()
        chart, use_markers = _make_chart(kind, sheet_label)

        src_ws = _resolve_source_ws(
            wb, ws_by_index, spec.get("source_sheet"), sheet_label, ws
        )
        data_ref = _ref_from_dict(src_ws, spec.get("data"), sheet_label, "data")
        titles_from_data = bool(spec.get("titles_from_data", True))

        if kind == "scatter":
            # 산점도는 x/y 를 별도 계열로 구성한다 — categories 가 x 축.
            cat_spec = spec.get("categories")
            if cat_spec is None:
                raise _err(
                    f"sheet {sheet_label!r}: scatter chart requires `categories` "
                    f"(the x-axis range).",
                    f"시트 {sheet_label!r}: scatter 차트에는 x축 범위인 "
                    f"`categories` 가 필요합니다.",
                )
            x_ref = _ref_from_dict(src_ws, cat_spec, sheet_label, "categories")
            series = Series(data_ref, x_ref, title_from_data=titles_from_data)
            chart.series.append(series)
        else:
            chart.add_data(data_ref, titles_from_data=titles_from_data)
            cat_spec = spec.get("categories")
            if cat_spec is not None:
                chart.set_categories(
                    _ref_from_dict(src_ws, cat_spec, sheet_label, "categories")
                )

        if use_markers:
            for s in chart.series:
                s.marker = Marker(symbol="circle", size=7)

        title = spec.get("title")
        if title:
            chart.title = str(title)
        # openpyxl 의 width/height 단위는 cm 다 (기본 15 x 7.5).
        for key, attr in (("width_cm", "width"), ("height_cm", "height")):
            if spec.get(key) is not None:
                try:
                    setattr(chart, attr, float(spec[key]))
                except (TypeError, ValueError) as exc:
                    raise _err(
                        f"sheet {sheet_label!r}: chart `{key}` must be a number.",
                        f"시트 {sheet_label!r}: 차트 `{key}` 는 숫자여야 합니다.",
                    ) from exc
        if spec.get("x_axis_title") and getattr(chart, "x_axis", None) is not None:
            chart.x_axis.title = str(spec["x_axis_title"])
        if spec.get("y_axis_title") and getattr(chart, "y_axis", None) is not None:
            chart.y_axis.title = str(spec["y_axis_title"])

        anchor = str(spec.get("anchor") or "A1").strip().upper()
        ws.add_chart(chart, anchor)


# ─────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────
def xlsx_from_sheets(sheets: list[dict[str, Any]]) -> bytes:
    """sheet spec → 스타일 잡힌 .xlsx bytes. 구조 오류는 ValueError."""
    if not isinstance(sheets, list) or not sheets:
        raise _err(
            "`sheets` must be a non-empty list.",
            "`sheets`는 비어있지 않은 리스트여야 합니다.",
        )

    wb = Workbook()
    wb.remove(wb.active)
    seen_titles: set[str] = set()
    ws_by_index: list[Any] = []

    # ── pass 1: 시트 생성 + 표/보고서 렌더 ──
    for idx, sheet in enumerate(sheets):
        if not isinstance(sheet, dict):
            raise _err(
                f"sheets[{idx}] must be a mapping.",
                f"sheets[{idx}] 항목은 객체(dict)여야 합니다.",
            )
        name = str(sheet.get("name") or "").strip()
        if not name:
            raise _err(
                f"sheets[{idx}]: every sheet needs a `name`.",
                f"sheets[{idx}]: 모든 시트에는 `name`이 필요합니다.",
            )
        title = name[:31]
        if title in seen_titles:
            raise _err(
                f"duplicate sheet name {title!r} (after 31-char truncation).",
                f"시트 이름 {title!r} 이 중복됩니다 (31자 절단 후).",
            )
        seen_titles.add(title)

        headers = sheet.get("headers")
        rows = sheet.get("rows")
        markdown = sheet.get("markdown")
        charts = sheet.get("charts")

        has_table = headers is not None or rows is not None
        has_markdown = markdown is not None
        has_charts = charts is not None
        if not (has_table or has_markdown or has_charts):
            raise _err(
                f"sheet {title!r}: needs at least one of `headers`+`rows`, "
                f"`markdown`, or `charts`.",
                f"시트 {title!r}: `headers`+`rows` / `markdown` / `charts` 중 "
                f"최소 하나가 필요합니다.",
            )

        ws = wb.create_sheet(title=title)
        ws_by_index.append(ws)

        if has_table:
            if not isinstance(headers, list) or not headers:
                raise _err(
                    f"sheet {title!r}: `headers` must be a non-empty list.",
                    f"시트 {title!r}: `headers`는 비어있지 않은 리스트여야 합니다.",
                )
            if not isinstance(rows, list):
                raise _err(
                    f"sheet {title!r}: `rows` must be a list.",
                    f"시트 {title!r}: `rows`는 리스트여야 합니다.",
                )
            n_cols = len(headers)

            # 헤더행
            for c, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=str(header))
                cell.font = Font(bold=True)
                cell.fill = _HEADER_FILL
                cell.border = _BORDER

            # 데이터행
            for r_idx, row in enumerate(rows):
                if not isinstance(row, list):
                    raise _err(
                        f"sheet {title!r}: rows[{r_idx}] must be a list.",
                        f"시트 {title!r}: rows[{r_idx}] 항목은 리스트여야 합니다.",
                    )
                if len(row) > n_cols:
                    raise _err(
                        f"sheet {title!r}: rows[{r_idx}] has {len(row)} cells "
                        f"but only {n_cols} headers.",
                        f"시트 {title!r}: rows[{r_idx}] 셀 수({len(row)})가 "
                        f"헤더 수({n_cols})를 초과합니다.",
                    )
                for c in range(n_cols):
                    value = row[c] if c < len(row) else ""
                    cell = ws.cell(row=r_idx + 2, column=c + 1, value=value)
                    cell.border = _BORDER

            # 열 폭: widths 지정 우선, 없으면 내용 기반 자동
            widths = sheet.get("widths")
            for c in range(1, n_cols + 1):
                letter = get_column_letter(c)
                if isinstance(widths, list) and c - 1 < len(widths):
                    try:
                        ws.column_dimensions[letter].width = max(
                            _MIN_WIDTH, min(_MAX_WIDTH, int(widths[c - 1]))
                        )
                        continue
                    except (TypeError, ValueError):
                        pass
                longest = max(
                    [len(str(headers[c - 1]))]
                    + [len(str(row[c - 1])) for row in rows if c - 1 < len(row)],
                    default=_MIN_WIDTH,
                )
                # CJK 는 2칸 폭이지만 근사만 — 정확한 폭은 편집 도구 몫
                ws.column_dimensions[letter].width = max(
                    _MIN_WIDTH, min(_MAX_WIDTH, longest + 2)
                )

            # 숫자 서식 (열 문자 기준, 데이터행에만)
            number_formats = sheet.get("number_formats")
            if isinstance(number_formats, dict):
                for letter, fmt in number_formats.items():
                    col_letter = str(letter).strip().upper()
                    for r in range(2, len(rows) + 2):
                        ws[f"{col_letter}{r}"].number_format = str(fmt)

        if has_markdown:
            if not isinstance(markdown, str):
                raise _err(
                    f"sheet {title!r}: `markdown` must be a string.",
                    f"시트 {title!r}: `markdown` 은 문자열이어야 합니다.",
                )
            if has_table:
                raise _err(
                    f"sheet {title!r}: `markdown` cannot be combined with "
                    f"`headers`/`rows` on the same sheet — split them into two sheets.",
                    f"시트 {title!r}: `markdown` 과 `headers`/`rows` 는 같은 시트에 "
                    f"함께 쓸 수 없습니다 — 시트를 분리하세요.",
                )
            _render_markdown_sheet(ws, markdown)

        # 틀고정: 명시값 우선. 표 시트는 기본 "A2", 그 외는 없음.
        if "freeze" in sheet:
            freeze = sheet.get("freeze")
            ws.freeze_panes = str(freeze) if freeze else None
        elif has_table:
            ws.freeze_panes = "A2"

    # ── pass 2: 차트 (모든 시트가 존재해야 source_sheet 참조가 성립) ──
    for idx, sheet in enumerate(sheets):
        charts = sheet.get("charts")
        if charts is None:
            continue
        title = str(sheet.get("name") or "").strip()[:31]
        _add_charts(wb, ws_by_index[idx], ws_by_index, charts, title)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
