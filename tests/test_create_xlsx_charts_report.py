"""create_document(.xlsx) — 차트 시트 · 보고서(markdown) 시트 렌더 검증.

v0.17 신규:
    - sheets[].charts   : 다른 시트의 셀 범위를 참조하는 살아있는 엑셀 차트
    - sheets[].markdown : markdown → 셀 + 서식 (보고서 시트)
    - sheets[].freeze   : 틀고정 명시 지정

회귀 축: **편집 왕복 후에도 차트/이미지가 보존**되어야 한다.
openpyxl 이 재저장 시 차트를 버린다는 통설이 있으나 3.1.5 기준으로는
보존된다 — 업그레이드로 퇴행하면 이 테스트가 잡는다.
"""
from __future__ import annotations

import io
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText

from document_adapter import create_document, load
from document_adapter.generate import xlsx_from_sheets

DATA_SHEET = {
    "name": "데이터",
    "headers": ["월", "매출", "비용"],
    "rows": [[f"{m}월", 100 + m * 10, 60 + m * 5] for m in range(1, 13)],
    "number_formats": {"B": "#,##0", "C": "#,##0"},
}

CHART_SHEET = {
    "name": "차트",
    "charts": [
        {
            "type": "column",
            "anchor": "B2",
            "title": "월별 매출/비용",
            "source_sheet": 0,
            "data": {"min_col": 2, "max_col": 3, "min_row": 1, "max_row": 13},
            "categories": {"min_col": 1, "max_col": 1, "min_row": 2, "max_row": 13},
            "titles_from_data": True,
            "width_cm": 20,
            "height_cm": 10,
            "x_axis_title": "월",
            "y_axis_title": "금액",
        }
    ],
}

REPORT_MD = (
    "# 상반기 분석\n\n"
    "본문에 **굵게** 와 *기울임* 과 `코드` 가 섞여 있습니다.\n\n"
    "## 발견\n\n"
    "- 첫 번째 항목\n"
    "- 두 번째 항목\n"
    "1. 번호 항목\n\n"
    "> 인용문입니다\n\n"
    "---\n\n"
    "| 분기 | 매출 |\n"
    "|---|---|\n"
    "| 1분기 | 350 |\n\n"
    "```\ncode line\n```\n"
)

REPORT_SHEET = {"name": "보고서", "markdown": REPORT_MD}


def _wb(sheets) -> "load_workbook":
    return load_workbook(io.BytesIO(xlsx_from_sheets(sheets)))


# ── 3시트 시나리오 ───────────────────────────────────────────────────
def test_three_sheet_scenario_structure():
    wb = _wb([DATA_SHEET, CHART_SHEET, REPORT_SHEET])
    assert wb.sheetnames == ["데이터", "차트", "보고서"]
    assert len(wb["차트"]._charts) == 1
    assert wb["데이터"].freeze_panes == "A2"
    # 보고서/차트 시트는 기본 틀고정 없음
    assert wb["보고서"].freeze_panes is None
    assert wb["차트"].freeze_panes is None


def test_chart_references_other_sheet():
    wb = _wb([DATA_SHEET, CHART_SHEET])
    chart = wb["차트"]._charts[0]
    refs = [str(s.val.numRef.f) for s in chart.series]
    assert len(refs) == 2, refs                      # 매출 + 비용
    assert all("데이터" in r for r in refs), refs     # 다른 시트를 참조
    assert chart.title is not None


def test_chart_xml_part_written():
    data = xlsx_from_sheets([DATA_SHEET, CHART_SHEET])
    names = zipfile.ZipFile(io.BytesIO(data)).namelist()
    assert any(n.startswith("xl/charts/chart") for n in names), names


def test_source_sheet_by_name():
    spec = dict(CHART_SHEET)
    spec["charts"] = [dict(CHART_SHEET["charts"][0], source_sheet="데이터")]
    wb = _wb([DATA_SHEET, spec])
    assert len(wb["차트"]._charts) == 1


def test_source_sheet_defaults_to_own_sheet():
    """source_sheet 생략 시 차트가 놓인 시트 자신을 참조한다."""
    sheet = {
        "name": "단일",
        "headers": ["항목", "값"],
        "rows": [["a", 1], ["b", 2]],
        "charts": [
            {
                "type": "pie",
                "anchor": "E2",
                "data": {"min_col": 2, "min_row": 1, "max_row": 3},
                "categories": {"min_col": 1, "min_row": 2, "max_row": 3},
            }
        ],
    }
    wb = _wb([sheet])
    assert len(wb["단일"]._charts) == 1


@pytest.mark.parametrize(
    "kind",
    [
        "column", "column_stacked", "bar", "bar_stacked",
        "line", "line_markers", "pie", "doughnut",
        "area", "area_stacked", "radar",
    ],
)
def test_every_chart_type_builds(kind):
    spec = {
        "name": "차트",
        "charts": [
            {
                "type": kind,
                "anchor": "B2",
                "source_sheet": 0,
                "data": {"min_col": 2, "min_row": 1, "max_row": 13},
                "categories": {"min_col": 1, "min_row": 2, "max_row": 13},
            }
        ],
    }
    wb = _wb([DATA_SHEET, spec])
    assert len(wb["차트"]._charts) == 1


def test_scatter_requires_categories():
    spec = {
        "name": "차트",
        "charts": [
            {
                "type": "scatter",
                "source_sheet": 0,
                "data": {"min_col": 2, "min_row": 1, "max_row": 13},
            }
        ],
    }
    with pytest.raises(ValueError) as exc:
        xlsx_from_sheets([DATA_SHEET, spec])
    assert "categories" in str(exc.value)


# ── 보고서(markdown) 시트 ────────────────────────────────────────────
def test_report_sheet_renders_all_block_kinds():
    ws = _wb([REPORT_SHEET])["보고서"]
    texts = [
        str(c.value) for row in ws.iter_rows() for c in row if c.value not in (None, "")
    ]
    joined = "\n".join(texts)
    assert "상반기 분석" in joined
    assert "발견" in joined
    assert "• 첫 번째 항목" in joined
    assert "인용문입니다" in joined
    assert "분기" in joined and "350" in joined      # 표
    assert "code line" in joined                      # 코드펜스


def test_report_heading_styles():
    ws = _wb([REPORT_SHEET])["보고서"]
    h1 = ws["A1"]
    assert h1.font.bold is True
    assert h1.font.size == 16
    assert h1.border.bottom.style == "medium"


def test_report_gridlines_hidden_and_widths():
    ws = _wb([REPORT_SHEET])["보고서"]
    assert ws.sheet_view.showGridLines is False
    assert ws.column_dimensions["A"].width == 3
    assert ws.column_dimensions["B"].width == 20


def test_report_inline_rich_text_preserved():
    """**굵게** / *기울임* / `코드` 가 run 단위로 살아있어야 한다.

    openpyxl 은 ``rich_text=True`` 로 열어야 CellRichText 를 돌려준다
    (기본값은 평문으로 평탄화).
    """
    data = xlsx_from_sheets([REPORT_SHEET])
    ws = load_workbook(io.BytesIO(data), rich_text=True)["보고서"]
    rich = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, CellRichText)
    ]
    assert rich, "no CellRichText cell found"
    flat = "".join(str(p) for r in rich for p in r)
    assert "굵게" in flat and "기울임" in flat


def test_rich_text_survives_adapter_roundtrip(tmp_path: Path):
    """어댑터 편집 왕복에서도 인라인 서식이 보존돼야 한다.

    XlsxAdapter._open() 이 ``rich_text=True`` 로 열지 않으면 재저장 시
    run 서식이 평문으로 뭉개진다 — 이 테스트가 그 회귀를 잡는다.
    """
    out = tmp_path / "rich.xlsx"
    create_document(out, sheets=[REPORT_SHEET])

    def rich_cells() -> int:
        wb = load_workbook(out, rich_text=True)
        return sum(
            1
            for row in wb["보고서"].iter_rows()
            for c in row
            if isinstance(c.value, CellRichText)
        )

    before = rich_cells()
    assert before >= 1

    doc = load(out)
    doc.save()          # 편집 없이 재저장만 해도 유실되면 안 된다
    doc.close()

    assert rich_cells() == before, "inline formatting lost on adapter re-save"


def test_report_bullet_is_indented_to_column_b():
    ws = _wb([REPORT_SHEET])["보고서"]
    hits = [
        c
        for row in ws.iter_rows()
        for c in row
        if c.value and str(c.value).startswith("• ")
    ]
    assert hits, "no bullet cell"
    assert all(c.column == 2 for c in hits)     # B 열
    assert all(c.alignment.indent == 1 for c in hits)


# ── freeze 옵션 ──────────────────────────────────────────────────────
def test_freeze_override_and_disable():
    a = dict(DATA_SHEET, name="고정", freeze="B3")
    b = dict(DATA_SHEET, name="해제", freeze="")
    wb = _wb([a, b])
    assert wb["고정"].freeze_panes == "B3"
    assert wb["해제"].freeze_panes is None


# ── 검증 실패(이중어) ────────────────────────────────────────────────
@pytest.mark.parametrize(
    "sheets, ko_fragment",
    [
        ([{"name": "빈시트"}], "최소 하나가 필요합니다"),
        (
            [dict(DATA_SHEET, markdown="# x")],
            "함께 쓸 수 없습니다",
        ),
        (
            [{"name": "c", "charts": [{"type": "nope", "data": {"min_col": 1, "min_row": 1}}]}],
            "지원하지 않는 차트",
        ),
        (
            [{"name": "c", "charts": [{"type": "column"}]}],
            "객체여야 합니다",
        ),
        (
            [
                DATA_SHEET,
                {
                    "name": "c",
                    "charts": [
                        {
                            "type": "column",
                            "source_sheet": 99,
                            "data": {"min_col": 1, "min_row": 1},
                        }
                    ],
                },
            ],
            "범위를 벗어났습니다",
        ),
        (
            [
                {
                    "name": "c",
                    "charts": [
                        {
                            "type": "column",
                            "data": {"min_col": 0, "min_row": 1},
                        }
                    ],
                }
            ],
            "올바르지 않습니다",
        ),
        ([{"name": "md", "markdown": 123}], "문자열이어야 합니다"),
        ([{"name": "c", "charts": "not-a-list"}], "리스트여야 합니다"),
    ],
)
def test_invalid_spec_raises_bilingual(sheets, ko_fragment):
    with pytest.raises(ValueError) as exc:
        xlsx_from_sheets(sheets)
    msg = str(exc.value)
    assert ko_fragment in msg, msg
    # 영어 문장도 함께 있어야 한다 (이중어 계약)
    assert any(c.isascii() and c.isalpha() for c in msg)


# ── 회귀: 편집 왕복 후 차트 보존 ─────────────────────────────────────
def test_charts_survive_edit_roundtrip(tmp_path: Path):
    out = tmp_path / "r.xlsx"
    create_document(out, sheets=[DATA_SHEET, CHART_SHEET, REPORT_SHEET])

    def chart_count() -> int:
        return len(load_workbook(out)["차트"]._charts)

    assert chart_count() == 1
    for i in range(3):
        wb = load_workbook(out)
        wb["데이터"][f"B{i + 2}"] = 999
        wb.save(out)
        assert chart_count() == 1, f"chart lost after round-trip #{i + 1}"


def test_charts_survive_adapter_edit(tmp_path: Path):
    """document_adapter 의 편집 API(set_cell) 경유 왕복에서도 보존."""
    out = tmp_path / "a.xlsx"
    create_document(out, sheets=[DATA_SHEET, CHART_SHEET])

    doc = load(out)
    doc.set_cell(0, 1, 1, "999")     # 데이터 시트(index 0) 수정
    doc.save()
    doc.close()

    assert len(load_workbook(out)["차트"]._charts) == 1


def test_generate_then_load_is_editable(tmp_path: Path):
    """생성-편집 왕복 계약: 차트/보고서 시트가 있어도 load() 가 성립한다."""
    out = tmp_path / "e.xlsx"
    create_document(out, sheets=[DATA_SHEET, CHART_SHEET, REPORT_SHEET])
    doc = load(out)
    schemas = doc.get_tables()
    names = [s.location for s in schemas]
    assert "데이터" in names
    doc.close()
