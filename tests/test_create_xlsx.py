"""create_document(.xlsx) — sheet spec 렌더 검증 + roundtrip."""
from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import load_workbook

from document_adapter import create_document, load
from document_adapter.generate import xlsx_from_sheets

SAMPLE_SHEETS = [
    {
        "name": "매출 요약",
        "headers": ["분기", "매출(억원)", "YoY"],
        "rows": [
            ["1분기", 120, "+12%"],
            ["2분기", 135.5, "+9%"],
            ["합계", "=SUM(B2:B3)", ""],
        ],
        "widths": [10, 14, 10],
        "number_formats": {"B": "#,##0"},
    },
    {
        "name": "기준정보",
        "headers": ["항목", "값"],
        "rows": [["기준일", "2026-07-09"]],
    },
]


def test_xlsx_render_types_and_styles() -> None:
    wb = load_workbook(io.BytesIO(xlsx_from_sheets(SAMPLE_SHEETS)))
    assert wb.sheetnames == ["매출 요약", "기준정보"]

    ws = wb["매출 요약"]
    # 헤더 스타일
    assert ws["A1"].value == "분기"
    assert ws["A1"].font.bold is True
    assert ws["A1"].fill.start_color.rgb.endswith("D9D9D9")
    # 숫자는 숫자 타입, 수식은 살아있는 수식
    assert ws["B2"].value == 120 and isinstance(ws["B2"].value, int)
    assert ws["B3"].value == 135.5
    assert ws["B4"].value == "=SUM(B2:B3)"
    # number_format (데이터행)
    assert ws["B2"].number_format == "#,##0"
    # 틀고정 + 열폭
    assert ws.freeze_panes == "A2"
    assert ws.column_dimensions["B"].width == 14


def test_auto_width_when_omitted() -> None:
    wb = load_workbook(io.BytesIO(xlsx_from_sheets(
        [{"name": "S", "headers": ["아주아주아주아주 긴 헤더"], "rows": []}]
    )))
    assert wb["S"].column_dimensions["A"].width >= 8


def test_short_rows_padded() -> None:
    wb = load_workbook(io.BytesIO(xlsx_from_sheets(
        [{"name": "S", "headers": ["a", "b", "c"], "rows": [["1"]]}]
    )))
    ws = wb["S"]
    assert ws["A2"].value == "1"
    assert ws["B2"].value in ("", None)


# -------- 검증 에러 (이중어 메시지) --------

@pytest.mark.parametrize(
    "sheets, ko_fragment",
    [
        ([], "리스트"),
        (["문자열"], "객체"),
        ([{"headers": ["a"], "rows": []}], "name"),
        (
            [{"name": "동일", "headers": ["a"], "rows": []},
             {"name": "동일", "headers": ["a"], "rows": []}],
            "중복",
        ),
        ([{"name": "S", "headers": [], "rows": []}], "headers"),
        ([{"name": "S", "headers": ["a"], "rows": "문자열"}], "rows"),
        ([{"name": "S", "headers": ["a"], "rows": ["행이아님"]}], "리스트"),
        ([{"name": "S", "headers": ["a"], "rows": [["1", "2"]]}], "초과"),
    ],
)
def test_invalid_spec_raises_bilingual(sheets, ko_fragment) -> None:
    with pytest.raises(ValueError) as exc:
        xlsx_from_sheets(sheets)
    assert ko_fragment in str(exc.value)


def test_roundtrip_generate_then_edit(tmp_path: Path) -> None:
    out = create_document(tmp_path / "매출.xlsx", sheets=SAMPLE_SHEETS)
    adapter = load(out)
    try:
        tables = adapter.get_tables(preview_rows=10)
        assert len(tables) == 2
        adapter.set_cell(0, 1, 2, "+13%")
        adapter.save()
    finally:
        adapter.close()

    wb = load_workbook(out)
    assert wb["매출 요약"]["C2"].value == "+13%"
