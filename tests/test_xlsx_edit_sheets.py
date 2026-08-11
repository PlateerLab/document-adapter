"""XLSX 편집 확장 (v0.19) — 행/열 삽입·삭제 + 시트 관리.

openpyxl 의 insert_rows/delete_rows/insert_cols/delete_cols 는 **셀 값만**
옮긴다. 병합 범위·행높이·열폭·삽입행 서식은 어댑터가 손으로 보정하므로,
이 테스트가 그 보정을 고정한다 (openpyxl 업그레이드 회귀 감지 포함).
"""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

from document_adapter import load
from document_adapter.base import CellOutOfBoundsError, NotImplementedForFormat


def _build(p: Path) -> None:
    """헤더 서식 + 세로병합(A4:A5) + 행높이(5행) + 열폭(C) + 차트."""
    wb = Workbook()
    ws = wb.active
    ws.title = "d"
    ws.append(["h1", "h2", "h3"])
    for i in range(1, 6):
        ws.append([f"r{i}", i * 10, i * 100])
    for c in range(1, 4):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).fill = PatternFill("solid", start_color="D9D9D9")
    ws.merge_cells("A4:A5")
    ws.row_dimensions[5].height = 40
    ws.column_dimensions["C"].width = 25
    ws2 = wb.create_sheet("c")
    ch = BarChart()
    ch.add_data(Reference(ws, min_col=2, min_row=1, max_row=6), titles_from_data=True)
    ws2.add_chart(ch, "B2")
    wb.save(p)


def _merges(ws) -> list[str]:
    return sorted(str(r) for r in ws.merged_cells.ranges)


@pytest.fixture()
def book(tmp_path: Path) -> Path:
    p = tmp_path / "b.xlsx"
    _build(p)
    return p


# ── insert_row ───────────────────────────────────────────────────────
def test_insert_row_shifts_values_merges_heights(book: Path):
    doc = load(book)
    idx = doc.insert_row(0, ["NEW", "1", "2"], at_row=1)
    doc.save()
    doc.close()
    assert idx == 1

    wb = load_workbook(book)
    ws = wb["d"]
    assert ws["A2"].value == "NEW"
    assert ws["A3"].value == "r1"          # 기존 행이 아래로 밀림
    assert "A5:A6" in _merges(ws)          # 병합이 한 칸 내려감
    assert ws.row_dimensions[6].height == 40   # 행높이도 따라 내려감
    assert len(wb["c"]._charts) == 1        # 차트 보존


def test_insert_row_inherits_adjacent_style(book: Path):
    """헤더 위 삽입이면 헤더 서식을 물려받는다."""
    doc = load(book)
    doc.insert_row(0, ["X", "9", "8"], at_row=0)
    doc.save()
    doc.close()

    ws = load_workbook(book)["d"]
    assert ws["A1"].value == "X"
    assert ws["A2"].value == "h1"
    assert ws["A1"].font.bold is True


def test_insert_row_appends_when_at_row_omitted(book: Path):
    doc = load(book)
    before = load_workbook(book)["d"].max_row
    idx = doc.insert_row(0, ["last", "1", "2"])
    doc.save()
    doc.close()
    assert idx == before
    ws = load_workbook(book)["d"]
    assert ws.cell(row=before + 1, column=1).value == "last"


def test_insert_row_numeric_values_become_numbers(book: Path):
    doc = load(book)
    doc.insert_row(0, ["n", "1234", "5,678"], at_row=1)
    doc.save()
    doc.close()
    ws = load_workbook(book)["d"]
    assert ws["B2"].value == 1234
    assert ws["C2"].value == 5678


# ── insert_column ────────────────────────────────────────────────────
def test_insert_column_shifts_values_and_widths(book: Path):
    doc = load(book)
    idx = doc.insert_column(0, ["new", "a", "b"], at_col=1)
    doc.save()
    doc.close()
    assert idx == 1

    ws = load_workbook(book)["d"]
    assert [ws.cell(row=1, column=c).value for c in (1, 2, 3)] == ["h1", "new", "h2"]
    assert ws.column_dimensions["D"].width == 25   # C 폭이 D 로 이동
    assert "A4:A5" in _merges(ws)                  # A 열 병합은 영향 없음


# ── delete_row / delete_column ───────────────────────────────────────
def test_delete_row_returns_values_and_shifts(book: Path):
    doc = load(book)
    removed = doc.delete_row(0, 1)
    doc.save()
    doc.close()

    assert removed[0] == "r1"
    ws = load_workbook(book)["d"]
    assert ws["A2"].value == "r2"
    assert "A3:A4" in _merges(ws)          # 병합이 한 칸 올라감
    assert ws.row_dimensions[4].height == 40


def test_delete_column_returns_values_and_shifts(book: Path):
    doc = load(book)
    removed = doc.delete_column(0, 1)
    doc.save()
    doc.close()

    assert removed[0] == "h2"
    ws = load_workbook(book)["d"]
    assert ws["B1"].value == "h3"
    assert ws.column_dimensions["B"].width == 25


# ── 병합 경계 가드 ───────────────────────────────────────────────────
def test_insert_row_crossing_merge_refused(book: Path):
    doc = load(book)
    with pytest.raises(NotImplementedForFormat) as exc:
        doc.insert_row(0, ["x"], at_row=4)     # A4:A5 한가운데
    doc.close()
    assert "merge" in str(exc.value)


def test_delete_row_crossing_merge_refused(book: Path):
    doc = load(book)
    with pytest.raises(NotImplementedForFormat):
        doc.delete_row(0, 3)                   # 병합 시작행
    doc.close()


def test_out_of_range_refused(book: Path):
    doc = load(book)
    with pytest.raises(CellOutOfBoundsError):
        doc.insert_row(0, ["x"], at_row=99)
    with pytest.raises(CellOutOfBoundsError):
        doc.delete_row(0, 99)
    with pytest.raises(CellOutOfBoundsError):
        doc.delete_column(0, 99)
    doc.close()


# ── 시트 관리 ────────────────────────────────────────────────────────
def test_add_rename_delete_sheet(book: Path):
    doc = load(book)
    idx = doc.add_sheet("요약")
    old = doc.rename_sheet(0, "원본데이터")
    gone = doc.delete_sheet(1)
    doc.save()
    doc.close()

    assert (idx, old, gone) == (2, "d", "c")
    assert load_workbook(book).sheetnames == ["원본데이터", "요약"]


def test_add_sheet_at_index(book: Path):
    doc = load(book)
    idx = doc.add_sheet("맨앞", at_index=0)
    doc.save()
    doc.close()
    assert idx == 0
    assert load_workbook(book).sheetnames[0] == "맨앞"


@pytest.mark.parametrize(
    "name, fragment",
    [
        ("", "비울 수 없습니다"),
        ("bad[name]", "사용할 수 없는"),
        ("d", "이미 존재"),
    ],
)
def test_invalid_sheet_name_refused(book: Path, name, fragment):
    doc = load(book)
    with pytest.raises(ValueError) as exc:
        doc.add_sheet(name)
    doc.close()
    assert fragment in str(exc.value)


def test_sheet_name_truncated_to_31(book: Path):
    doc = load(book)
    doc.add_sheet("가" * 40)
    doc.save()
    doc.close()
    assert any(len(t) == 31 for t in load_workbook(book).sheetnames)


def test_cannot_delete_last_sheet(book: Path):
    doc = load(book)
    doc.delete_sheet(1)
    with pytest.raises(ValueError) as exc:
        doc.delete_sheet(0)
    doc.close()
    assert "마지막" in str(exc.value)


# ── 수식 경고 판단 ───────────────────────────────────────────────────
def test_has_formulas_detects(tmp_path: Path):
    p = tmp_path / "f.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["a", "b"])
    ws.append([1, "=A2*2"])
    wb.create_sheet("plain").append(["x"])
    wb.save(p)

    doc = load(p)
    assert doc.has_formulas(0) is True
    assert doc.has_formulas(1) is False
    doc.close()


# ── 다른 포맷은 미지원이어야 한다 ────────────────────────────────────
def test_docx_still_refuses_sheet_ops(tmp_path: Path):
    from document_adapter import create_document

    p = tmp_path / "x.docx"
    create_document(p, markdown="# t\n\n| a | b |\n|---|---|\n| 1 | 2 |\n")
    doc = load(p)
    for fn in (
        lambda: doc.add_sheet("s"),
        lambda: doc.rename_sheet(0, "s"),
        lambda: doc.delete_sheet(0),
        lambda: doc.delete_row(0, 0),
        lambda: doc.delete_column(0, 0),
    ):
        with pytest.raises(NotImplementedForFormat):
            fn()
    doc.close()
